import re
import os
import json
import openpyxl 
from openpyxl import Workbook

class SysConfig:
    def __init__(self, file_name, system_name):
        with open(file_name, "r") as f:
            self.conf = json.load(f)["systems"]
            self.conf = self.conf[system_name]

def create_sheet(write_wb):
    write_ws = write_wb.create_sheet("파일 리스트")
    write_ws.append(["path", "filename", "ext", "filetype"])
    return write_ws

sys_config = SysConfig("./config.json", "hnb-sc-batch")
print(sys_config.conf["div_name"], sys_config.conf["system_name"])

# Create or Load Excel File
if os.path.isfile(sys_config.conf["save_file_name"]):
    os.remove(sys_config.conf["save_file_name"])
try :
    write_wb = openpyxl.load_workbook(sys_config.conf["save_file_name"])
except FileNotFoundError:
    write_wb = Workbook()

# Create or Load Excel Sheet
try :
    write_ws = write_wb["file_list"]
except KeyError:
    write_ws = create_sheet(write_wb)

print('except file list')
succ_cnt = 1;skip_cnt = 0
for (path, dir, files) in os.walk(sys_config.conf["system_root"]):
    for file_name in files:
        file_ext  = os.path.splitext(file_name)[-1]
        file_path = re.sub(r"\\","/",path)
        
        # Except file process
        except_code = "0000"
        if re.search("/WEB-INF/",file_path) is not None:
            except_code = "D001"
        elif re.search("/\.svn/",file_path) is not None:
            except_code = "D002"
        elif re.search("mq0adm",file_path) is not None:
            except_code = "D002"
        elif re.search("/(old|bac?k|upload|log|te?mp|bak|work|unused|backup)$",file_path, flags=re.IGNORECASE) is not None:
            except_code = "D011"
        elif re.search("/(old|bac?k|upload|log|te?mp|bak|work|unused|backup)/",file_path, flags=re.IGNORECASE) is not None:
            except_code = "D012"
        elif file_ext == ".class":
            except_code = "E001"
        elif file_ext == ".log":
            except_code = "E002"
        elif re.search("\.(\d{6,8}|bac?k|old).{0,5}$",file_name) is not None:
            except_code = "F001"
        elif re.search("_(\d{6,8}|imsi|TEST|tmp).{0,5}$",file_name) is not None:
            except_code = "F002"

        if except_code == "D001" :
            skip_cnt = skip_cnt + 1
            continue
        if except_code == "D002" :
            skip_cnt = skip_cnt + 1
            continue
        if except_code != "0000" :
            skip_cnt = skip_cnt + 1
            #print(f"[{except_code}] {file_path}/{file_name}")
            continue

        # Choose File type
        file_type = "etc"
        if re.search("online", sys_config.conf["save_file_name"]) is not None:
            if re.search(sys_config.conf["sql_root"],file_path) is not None:
                file_type = "sql"
            elif re.search(sys_config.conf["was_root"],file_path) is not None:
                file_type = "was"
            elif re.search(sys_config.conf["mip_root"],file_path) is not None:
                file_type = "mip"
            elif re.search(sys_config.conf["web_root"],file_path) is not None:
                if file_ext == ".xml" :
                    file_type = "web"
            elif re.search(sys_config.conf["ozr_root"],file_path) is not None:
                file_type = "ozr"
        elif re.search("batch", sys_config.conf["save_file_name"]) is not None:
            file_type = file_ext
        
        # Write excel rows
        succ_cnt = succ_cnt + 1
        link = f'=HYPERLINK(A{succ_cnt}&"/"&B{succ_cnt},"파일열기")'
        write_ws.append([file_path, file_name, file_ext, file_type, link])
        
    write_wb.save(sys_config.conf["save_file_name"])

print('\n\n')
print(f'{sys_config.conf["save_file_name"]} exported...')
print(f'succ:[{succ_cnt:05d}] / skip:[{skip_cnt:05d}]')
