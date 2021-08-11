import openpyxl 
from openpyxl import Workbook
import xml.etree.ElementTree as xml
import re
import os

def remove_comment(text, text_type):
    #범위주석제거
    text = re.sub(r"/\*([^*]|[\r\n]|(\*+([^*/]|[\r\n])))*\*+/", "", text)
    
    #한줄주석제거
    text = re.sub(r"//.*[\r\n]", "\n", text)
    
    #빈줄제거
    text = re.sub(r"\n\s*\n", "\n", text)
    
    return text

class SysConfig:
    def __init__(self):
        self.set_cvs_hq()

    def set_hnb_hq(self):
        self.system_root = "D:/Tools/NCDStudio/workspace/hbhq/trunk/apps/hbhq/"
        self.web_root = self.system_root + "web/mip/CSHQ/"
        self.ozr_root = self.system_root + "web/ozr/"
        self.mip_root = self.system_root + "devonhome/navigation/mip/cshq/"
        self.was_root = self.system_root + "src/"
        self.sql_root = self.system_root + "devonhome/xmlquery/"
        self.class_root = self.system_root + "web/WEB-INF/"

    def set_cvs_hq(self):
        self.system_root = "D:/Tools/NCDStudio/workspace/cshq/apps/cshq/"
        self.web_root = self.system_root + "web/mip/CSHQ/"
        self.ozr_root = self.system_root + "web/ozr/"
        self.mip_root = self.system_root + "devonhome/navigation/mip/cshq/"
        self.was_root = self.system_root + "src/"
        self.sql_root = self.system_root + "devonhome/xmlquery/"
        self.class_root = self.system_root + "web/WEB-INF/"

class FileController:
    def __init__(self, root_dir, file_name, file_type):
        self.__init__(root_dir, file_name + "." + file_type)
        
    def __init__(self, root_dir, file_name):
        self.root_dir  = root_dir
        tmp = re.search(r"/?((\S*/)*)(\S+)\.(\S+)", file_name)
        try:
            self.local_dir = tmp.group(1)
            self.file_name = tmp.group(3)
            self.file_type = tmp.group(4)
        except:
            print(file_name)

    #def __del__(self):
        #print("Destroy")

    def getFullFileName(self):
        return self.root_dir + self.local_dir + self.file_name + "." + self.file_type

    def getFileName(self):
        return "/" + self.local_dir + self.file_name 

    def read(self):
        self.read_name = "read_" + self.file_type
        self.read_func = getattr(self, self.read_name, lambda:"default")
        return self.read_func()

    def read_xml(self):
        return ""
    
    def read_java(self):
        return ""

def create_sheet(write_wb):
    write_ws = write_wb.create_sheet("Breadcrumb")
    write_ws.append(["web","","","was","","","sql",""])
    write_ws.append(["web (xml)","mip (xml)","key","cmd (java)","biz (java)","key","dao (xml)","key"])
    return write_ws

sys_config = SysConfig()
save_file_name = "D:\\asdf.xlsx"

# Create or Load Excel File
if os.path.isfile(save_file_name):
    os.remove(save_file_name)

try :
    write_wb = openpyxl.load_workbook(save_file_name)
except FileNotFoundError:
    write_wb = Workbook()

# Create or Load Excel Sheet
try :
    write_ws = write_wb.get_sheet_by_name("Breadcrumb")
except KeyError:
    write_ws = create_sheet(write_wb)

#print(sys_config.web_root)
#for (path, dir, files) in os.walk(sys_config.web_root + "cm_cometc/"):
for (path, dir, files) in os.walk(sys_config.web_root):
    for filename in files:
        ext = os.path.splitext(filename)[-1]
        if (ext == ".xml"):
            web_file = path + "/" + filename
            #web_file = web_file[63:] # 랄라블라
            web_file = web_file[57:] # 편의점
            web = FileController(sys_config.web_root, web_file)
            #print("web :", web.getFileName())

            mips = ""
            try :
                tree = xml.parse(web.getFullFileName())
                root = tree.getroot()
                leaf = root.find("Script").text
                leaf = remove_comment(leaf, "java")
                #print(leaf)
                mips = re.findall(r'\"\S+\.mip\"', leaf)
            except:
                print("Fail : web file compile.", web.getFileName())
                write_ws.append([web.getFileName()])
                #write_ws.append(web_file)

            if (mips == ""): continue

            for m in mips:

                mip = ""
                try:
                    # web 파일에서 참조하는 mip 파일명을 추출한다. 
                    tmp = re.search(r"(/.*)/(\S+)\.mip", m)
                    if tmp is None:
                        raise Exception('MIP 파일 추출 실패')

                    # mip 파일을 읽어서 cmd 파일명을 추출한다.
                    mip = FileController(sys_config.mip_root, tmp.group(1) + ".xml")
                    #print("mip :", mip.getFileName())
                    if os.path.isfile(mip.getFullFileName()):
                        tmp_tree = xml.parse(mip.getFullFileName())
                    else:
                        raise Exception('파일 없음')
                except:
                    print("Fail : mip file compile.", m)
                    #print(filename, m, tmp)
                    write_ws.append([web.getFileName(), m])
                    continue

                tmp2 = ""
                try:
                    web_key = tmp.group(2)
                    query = 'action[@name="'+web_key+'"]'
                    tmp2 = tmp_tree.find(query).find("command").text.strip()
                    #print("name:", tmp2.group(2))
                except:
                    print("Fail : cmd file compile.")
                    write_ws.append([web.getFileName(), mip.getFileName(), web_key])
                    continue

                java = ""
                try:
                    # cmd 파일을 open 하여 java 소스를 불러온다.
                    cmd = FileController(sys_config.was_root, re.sub(r"\.", "/", tmp2) +".java")
                    #print("cmd :", cmd.getFileName())

                    f = open(cmd.getFullFileName(),"r",encoding="utf-8")
                    java = remove_comment(f.read(), "java")
                    f.close()
                except:
                    print("Fail : cmd file open.", cmd.getFileName())
                    write_ws.append([web.getFileName(), mip.getFileName(), web_key])

                if (java == ""): continue

                
                #print(java)
                biz_func_nm = ""
                try:
                    # cmd 에서 호출하는 biz 함수명을 추출한다.
                    biz_func_nm = re.search(r"biz\.(.+)[(]",java).group(1)
                    biz_class_nm = re.search(r"(\S+)\s+biz",java).group(1)
                    #print("biz_func_nm:", biz_func_nm)
                    #print("biz_class_nm:", biz_class_nm)
                    
                    # 현재 cmd 패키지 경로로 biz 패키지 경로를 추출한다.
                    #tmp = re.search(r"import (.*biz.*);",java).group(1)
                    #print("import (.*biz\."+biz_class_nm+");", tmp)
                    tmp = re.search("import (.*biz\."+biz_class_nm+");",java).group(1)

                    # biz 파일을 open 하여 java 소스를 불러온다.
                    biz = FileController(sys_config.was_root, re.sub(r"\.", "/", tmp) +".java")
                    #print("biz :", biz.getFileName())
                    f = open(biz.getFullFileName(),"r",encoding="utf-8")
                    java = remove_comment(f.read(), "java")
                    f.close()
                except:
                    print("Fail : cmd file compile.", cmd.getFileName())
                    write_ws.append([web.getFileName(), mip.getFileName(), web_key, cmd.getFileName()])
                    continue

                try :
                    pattern = biz_func_nm + ".*?\(\"(/.+?)/(\w+)\""
                    java = re.search(pattern, java, flags=re.DOTALL)
                    #print(java.group(0))
                    sql = FileController(sys_config.sql_root, java.group(1) +".xml")
                    sql_key = java.group(2)
                    write_ws.append([web.getFileName(), mip.getFileName(), web_key, cmd.getFileName(), biz.getFileName(), biz_func_nm, sql.getFileName(), sql_key])
                except:
                    print("Fail : biz file compile.")
                    write_ws.append([web.getFileName(), mip.getFileName(), web_key, cmd.getFileName(), biz.getFileName(), biz_func_nm])
                #print("sql :", sql.getFileName())
                #print("dao :", java.group(2))

                

                #print("==========================================")

    write_wb.save(save_file_name)

'''
for x in leaf:
    print(x.text)
'''
'''
'x in WEB 소스파일 리스트
    'content = 파일파싱(x, "xml")
    'mip[] = mip추출(content, "script")
    'm in mip
        'c = cmd추출(m, "xml")
        'b = biz추출(c, "java")
        'q = query추출(b, "java")
'''

