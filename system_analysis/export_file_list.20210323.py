import openpyxl 
from openpyxl import Workbook
import xml.etree.ElementTree as xml
import re
import os
import json

class SysConfig:
    def __init__(self, file_name, system_name):
        with open(file_name, "r") as f:
            self.conf = json.load(f)["systems"]
            self.conf = self.conf[system_name]

#class SysConfig:
#    def __init__(self):
#        self.set_cvs_hq()
#
#    def set_hnb_hq(self):
#        self.system_root = "D:/Tools/NCDStudio/workspace/hbhq/trunk/apps/hbhq/"
#        self.web_root = self.system_root + "web/mip/CSHQ/"
#        self.ozr_root = self.system_root + "web/ozr/"
#        self.mip_root = self.system_root + "devonhome/navigation/mip/cshq/"
#        self.was_root = self.system_root + "src/"
#        self.sql_root = self.system_root + "devonhome/xmlquery/"
#        self.class_root = self.system_root + "web/WEB-INF/"
#
#    def set_cvs_hq(self):
#        self.system_root = "D:/Tools/NCDStudio/workspace/cshq/apps/cshq/"
#        self.web_root = self.system_root + "web/mip/CSHQ/"
#        self.ozr_root = self.system_root + "web/ozr/"
#        self.mip_root = self.system_root + "devonhome/navigation/mip/cshq/"
#        self.was_root = self.system_root + "src/"
#        self.sql_root = self.system_root + "devonhome/xmlquery/"
#        self.class_root = self.system_root + "web/WEB-INF/"

def create_sheet(write_wb):
    write_ws = write_wb.create_sheet("파일 리스트")
    write_ws.append(["path", "filename", "ext", "filetype"])
    return write_ws

sys_config = SysConfig("./config.json", "hnb-online")
save_file_name = "D:\\affff.xlsx"

# Create or Load Excel File
os.remove(save_file_name)
try :
    write_wb = openpyxl.load_workbook(save_file_name)
except FileNotFoundError:
    write_wb = Workbook()

# Create or Load Excel Sheet
try :
    write_ws = write_wb.get_sheet_by_name("파일 리스트")
except KeyError:
    write_ws = create_sheet(write_wb)

row = 1
for (path, dir, files) in os.walk(sys_config.system_root):
    for file_name in files:
        file_ext  = os.path.splitext(file_name)[-1]
        file_path = re.sub(r"\\","/",path)
        
        # Except class files
        if file_ext == ".class":
            continue

        # Choose File type
        file_type = "etc"
        if re.match(sys_config.sql_root,file_path) is not None:
            file_type = "sql"
        elif re.match(sys_config.was_root,file_path) is not None:
            file_type = "was"
        elif re.match(sys_config.mip_root,file_path) is not None:
            file_type = "mip"
        elif re.match(sys_config.web_root,file_path) is not None:
            if file_ext == ".xml" :
                file_type = "web"
        elif re.match(sys_config.ozr_root,file_path) is not None:
            file_type = "ozr"

        row = row + 1
        link = f'=HYPERLINK(A{row}&"/"&B{row},"파일열기")'
        write_ws.append([file_path, file_name, file_ext, file_type, link])
            
write_wb.save(save_file_name)
